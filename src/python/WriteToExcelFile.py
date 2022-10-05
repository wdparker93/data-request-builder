import openpyxl, os
from pathlib import Path
from flask import Flask
from flask import request, make_response
from flask_cors import CORS

app = Flask(__name__)

CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

@app.route("/writeToExcelFile", methods=['POST'])
def writeToExcelFile():
    
    tableNameDict = request.json['tableNames']
    fieldNameDict = request.json['fieldNames']
    checkBoxDict = request.json['checkBoxes']
    print(tableNameDict)
    print(fieldNameDict)
    tableCheckBoxDict = buildTableCheckBoxDict(checkBoxDict)
    fieldCheckBoxDict = buildFieldCheckBoxDict(checkBoxDict)
    print(tableCheckBoxDict)
    print(fieldCheckBoxDict)

    #Get to the project root folder,
    #then navigate to the output excel file
    currentDirectory = Path(os.getcwd())
    parentDirectory = currentDirectory
    for i in range(2):
        parentDirectory = parentDirectory.parent
    os.chdir(str(parentDirectory) + '\\docs\\output')
    outputPath = os.getcwd() + '\\OutputDataDefinition.xlsx'

    outputDataDefWB = openpyxl.load_workbook(outputPath)
    tableSheet = outputDataDefWB['Table_Names']
    fieldSheet = outputDataDefWB['Field_Names']

    nameEntryCol = 1
    outputRow = 1

    with(open(outputPath, "a")) as f:
        #Clear the previous workbook's contents
        for row in tableSheet['A1':'A1000']:
            for cell in row:
                cell.value = None
        for row in fieldSheet['A1':'A1000']:
            for cell in row:
                cell.value = None
        #Enter table names to table sheet
        for key in tableNameDict:
            if tableCheckBoxDict[key]:
                tableSheet.cell(row=outputRow, column=nameEntryCol).value = tableNameDict[key]
                outputRow += 1

        
        #Reset the output row to row 1 before write to fields sheet
        outputRow = 1

        #Enter field names in field sheet
        for key in fieldNameDict:
            if tableCheckBoxDict[key]:
                fieldNames = fieldNameDict[key]
                for i in range(len(fieldNames)):
                    if fieldCheckBoxDict[key][i]:
                        fieldSheet.cell(row=outputRow, column=nameEntryCol).value = fieldNames[i]
                        outputRow += 1

        outputDataDefWB.save(filename=os.getcwd() + '\\OutputDataDefinition.xlsx')
    
    return make_response("Success", 200)

#Converts the checkBoxDict to a format that matches the table name dictionary
def buildTableCheckBoxDict(checkBoxParam):
    returnDictionary = {}
    for i in range(len(checkBoxParam)):
        checkBoxEntry = checkBoxParam[i]
        checkBoxId = checkBoxEntry[0]
        checkBoxCheckedVal = checkBoxEntry[1]
        if "field" not in checkBoxId:
            #Get the tableNumber which will be the key in the returnDictionary
            tableNumStart = checkBoxId.index("table-") + 6
            tableNumEnd = checkBoxId.index("-name")
            tableNumber = int(checkBoxId[tableNumStart:tableNumEnd])
            tableNumberKey = str(tableNumber)
            #Add the fields to the dicionary's entry's array
            returnDictionary[tableNumberKey] = checkBoxCheckedVal
    return returnDictionary


#Converts the checkBoxDict to a format that matches the field name dictionary
def buildFieldCheckBoxDict(checkBoxParam):
    returnDictionary = {}
    for i in range(len(checkBoxParam)):
        checkBoxEntry = checkBoxParam[i]
        checkBoxId = checkBoxEntry[0]
        checkBoxCheckedVal = checkBoxEntry[1]
        if "field" in checkBoxId:
            #Get the tableNumber which will be the key in the returnDictionary
            tableNumStart = checkBoxId.index("table-") + 6
            tableNumEnd = checkBoxId.index("-field")
            tableNumber = int(checkBoxId[tableNumStart:tableNumEnd])
            tableNumberKey = str(tableNumber)
            #Get the fieldNumber which will be used as indexes in the returnDictionary
            fieldNumStart = checkBoxId.index("field-") + 6
            fieldNumEnd = checkBoxId.index("-name")
            fieldNumber = int(checkBoxId[fieldNumStart:fieldNumEnd])
            fieldIndex = fieldNumber - 1
            #Add the fields to the dicionary's entry's array
            if tableNumberKey in returnDictionary:
                returnDictionary[tableNumberKey].insert(fieldIndex, checkBoxCheckedVal)
            else:
                returnDictionary[tableNumberKey] = []
                returnDictionary[tableNumberKey].insert(0, checkBoxCheckedVal)
    return returnDictionary

if __name__ == "__main__":
    app.run()