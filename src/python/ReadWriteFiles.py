import openpyxl, os
from pathlib import Path
from flask import Flask
from flask import request, make_response
from flask_cors import CORS

app = Flask(__name__)

CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

@app.route("/saveStateToTextFile", methods=['POST'])
def saveToTextFile():
    tableNameDict = request.json['tableNames']
    fieldNameDict = request.json['fieldNames']
    checkBoxDict = request.json['checkBoxes']
    tableCheckBoxDict = buildTableCheckBoxDict(checkBoxDict)
    fieldCheckBoxDict = buildFieldCheckBoxDict(checkBoxDict)

    #Get to the project root folder,
    #then navigate to the output excel file
    currentDirectory = Path(os.getcwd())
    parentDirectory = currentDirectory
    for i in range(2):
        parentDirectory = parentDirectory.parent
    os.chdir(str(parentDirectory) + '\\docs\\output')
    outputPath = os.getcwd() + '\\SessionConfigOutput.txt'

    #Open and close file to clear contents before write new config
    open(outputPath, 'w').close()

    with(open(outputPath, 'w')) as f:
        for key in tableNameDict:
            f.write(tableNameDict[key])
            f.write(':')
            tableCheckValue = 'true'
            if (tableCheckBoxDict[key] == False):
                tableCheckValue = 'false'
            f.write(tableCheckValue)
            f.write('\n')
            fieldNameArr = fieldNameDict[key]
            fieldBoxArr = fieldCheckBoxDict[key]
            for i in range (len(fieldNameArr)):
                f.write(fieldNameArr[i])
                f.write(':')
                fieldCheckValue = 'true'
                if (fieldBoxArr[i] == False):
                    fieldCheckValue = 'false'
                f.write(fieldCheckValue)
                if i < (len(fieldNameArr) - 1):
                    f.write(',')
            f.write('\n')

    return make_response("Success", 200)

@app.route("/writeToExcelFile", methods=['POST'])
def writeToExcelFile():
    tableNameDict = request.json['tableNames']
    fieldNameDict = request.json['fieldNames']
    checkBoxDict = request.json['checkBoxes']
    tableCheckBoxDict = buildTableCheckBoxDict(checkBoxDict)
    fieldCheckBoxDict = buildFieldCheckBoxDict(checkBoxDict)

    #Get to the project root folder,
    #then navigate to the output excel file
    currentDirectory = Path(os.getcwd())
    parentDirectory = currentDirectory
    for i in range(2):
        parentDirectory = parentDirectory.parent
    os.chdir(str(parentDirectory) + '\\docs\\output')
    outputPath = os.getcwd() + '\\OutputDataDefinition.xlsx'

    outputDataDefWB = openpyxl.load_workbook(outputPath)
    tableSheet = outputDataDefWB['Table_Names']
    fieldSheet = outputDataDefWB['Field_Names']

    nameEntryCol = 1
    outputRow = 1

    with(open(outputPath, "a")) as f:
        #Clear the previous workbook's contents
        for row in tableSheet['A1':'A1000']:
            for cell in row:
                cell.value = None
        for row in fieldSheet['A1':'A1000']:
            for cell in row:
                cell.value = None
        #Enter table names to table sheet
        for key in tableNameDict:
            if tableCheckBoxDict[key]:
                tableSheet.cell(row=outputRow, column=nameEntryCol).value = tableNameDict[key]
                outputRow += 1

        
        #Reset the output row to row 1 before write to fields sheet
        outputRow = 1

        #Enter field names in field sheet
        for key in fieldNameDict:
            if tableCheckBoxDict[key]:
                fieldNames = fieldNameDict[key]
                for i in range(len(fieldNames)):
                    if fieldCheckBoxDict[key][i]:
                        fieldSheet.cell(row=outputRow, column=nameEntryCol).value = fieldNames[i]
                        outputRow += 1

        outputDataDefWB.save(filename=os.getcwd() + '\\OutputDataDefinition.xlsx')
    
    return make_response("Success", 200)

@app.route("/readFromTextFile", methods=['POST'])
def readFromTextFile():
    #Get file stream from SpooledTemporaryFile
    storage = request.files['files']
    inputText = ''
    with storage.stream as f:
        inputText = f.read()
    #Write byte stream to temporary text file to make it easier when parsing
    currentDirectory = os.getcwd()
    tempTextFile = currentDirectory + '\\TempOutputFile.txt'
    with open(tempTextFile, 'wb') as f:
        f.write(inputText)
    #Parse temporary text file into data structures
    tableDict = {}
    tableCheckBoxDict = {}
    fieldNameDict = {}
    fieldCheckBoxDict = {}
    with open(tempTextFile, 'r') as f:
        lineNum = 1
        tableCounter = 1
        for line in f.readlines():
            line = line.replace('\r', '')
            line = line.replace('\n', '')
            if lineNum % 2 != 0:
                tableCheckBoxPresent = True
                indexOfColon = 0
                tableName = ''
                checkBoxValue = ''
                try:
                    indexOfColon = line.index(':')
                except:
                    tableCheckBoxPresent = False
                    pass
                if tableCheckBoxPresent == True:
                    tableName = line[:indexOfColon]
                    checkBoxValue = line[indexOfColon + 1:]
                else:
                    tableName = line
                    checkBoxValue = 'false'
                tableDict[str(tableCounter)] = tableName
                tableCheckBoxDict[str(tableCounter)] = checkBoxValue
            else:
                fieldArr = [line.split(',')]
                fieldArr = fieldArr[0]
                fieldNameArr = []
                checkBoxValueArr = []
                for field in fieldArr:
                    fieldCheckBoxPresent = True
                    indexOfColon = 0
                    fieldName = ''
                    checkBoxValue = ''
                    try:
                        indexOfColon = field.index(':')
                    except:
                        fieldCheckBoxPresent = False
                        pass
                    if fieldCheckBoxPresent == True:
                        fieldName = field[:indexOfColon]
                        checkBoxValue = field[indexOfColon + 1:]
                    else:
                        fieldName = field
                        checkBoxValue = 'false'
                    fieldNameArr.append(fieldName)
                    checkBoxValueArr.append(checkBoxValue)
                fieldNameDict[str(tableCounter)] = fieldNameArr
                fieldCheckBoxDict[str(tableCounter)] = checkBoxValueArr
                tableCounter += 1
            lineNum += 1
    returnDict = {}
    returnDict['Tables'] = tableDict
    returnDict['Table-Boxes'] = tableCheckBoxDict
    returnDict['Fields'] = fieldNameDict
    returnDict['Field-Boxes'] = fieldCheckBoxDict
    #Delete the temporary text file
    os.remove(tempTextFile)
    #Return the data structure with fields assigned and organized
    return make_response(returnDict, 200)

#Converts the checkBoxDict to a format that matches the table name dictionary
def buildTableCheckBoxDict(checkBoxParam):
    returnDictionary = {}
    for i in range(len(checkBoxParam)):
        checkBoxEntry = checkBoxParam[i]
        checkBoxId = checkBoxEntry[0]
        checkBoxCheckedVal = checkBoxEntry[1]
        if "field" not in checkBoxId:
            #Get the tableNumber which will be the key in the returnDictionary
            tableNumStart = checkBoxId.index("table-") + 6
            tableNumEnd = checkBoxId.index("-name")
            tableNumber = int(checkBoxId[tableNumStart:tableNumEnd])
            tableNumberKey = str(tableNumber)
            #Add the fields to the dicionary's entry's array
            returnDictionary[tableNumberKey] = checkBoxCheckedVal
    return returnDictionary

#Converts the checkBoxDict to a format that matches the field name dictionary
def buildFieldCheckBoxDict(checkBoxParam):
    returnDictionary = {}
    for i in range(len(checkBoxParam)):
        checkBoxEntry = checkBoxParam[i]
        checkBoxId = checkBoxEntry[0]
        checkBoxCheckedVal = checkBoxEntry[1]
        if "field" in checkBoxId:
            #Get the tableNumber which will be the key in the returnDictionary
            tableNumStart = checkBoxId.index("table-") + 6
            tableNumEnd = checkBoxId.index("-field")
            tableNumber = int(checkBoxId[tableNumStart:tableNumEnd])
            tableNumberKey = str(tableNumber)
            #Get the fieldNumber which will be used as indexes in the returnDictionary
            fieldNumStart = checkBoxId.index("field-") + 6
            fieldNumEnd = checkBoxId.index("-name")
            fieldNumber = int(checkBoxId[fieldNumStart:fieldNumEnd])
            fieldIndex = fieldNumber - 1
            #Add the fields to the dicionary's entry's array
            if tableNumberKey in returnDictionary:
                returnDictionary[tableNumberKey].insert(fieldIndex, checkBoxCheckedVal)
            else:
                returnDictionary[tableNumberKey] = []
                returnDictionary[tableNumberKey].insert(0, checkBoxCheckedVal)
    return returnDictionary

if __name__ == "__main__":
    app.run()